# *** PRE AND POST PROCESSING XLSX CONTENT *** #
# Automatically changing the format of XLSX files for translation

import os
import openpyxl

from pathlib import Path
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList
from openpyxl.utils import get_column_letter

sample_xlsx_dir = 'files'
sample_xlsx_file = 'DPDHL-Chatbot-Export_pre_processed.xlsx'
sample_xlsx_path = os.path.join(sample_xlsx_dir, sample_xlsx_file)


# *** CHANGE CELL BACKGROUND COLOR *** #
# Change the background color of cell ranges in XLSX files

def toggle_background_color(xlsx_path, color=True, out_file_suffix='_bg_color_changed'):
    """
    Changes the background color of cells in the first row of an XLSX document
    :param xlsx_path: path of XLSX file to be formatted
    :param color: if True, cells will be colored red. If False, cell color will be removed
    :param out_file_suffix: Suffix added to the output file before saving
    :return: None
    """
    workbook = openpyxl.load_workbook(xlsx_path)
    for worksheet in workbook:
        column = worksheet['A']
        for cell in column:
            if color:
                cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
            else:
                cell.fill = PatternFill(fill_type=None)
    hidden_out_dir = os.path.dirname(xlsx_path)
    hidden_out_file = Path(xlsx_path).stem + out_file_suffix + '.xlsx'
    hidden_out_path = os.path.join(hidden_out_dir, hidden_out_file)
    workbook.save(hidden_out_path)


toggle_background_color(sample_xlsx_path)


# *** CHANGE CELL BACKGROUND COLOR BY VALUE *** #
# Change the background color of non translatable cell ranges in XLSX files

# TODO: We can now color set cell ranges, but what if we only want to color certain cells?
# TODO: For example, if the value in a column "Translate" is False, we can mark the entire row
# TODO: We'd like to write a function "toggle_non_translatable_background_color"
# TODO: The function should take the following arguments:
# TODO: - xlsx_path: positional argument, path of the XLSX file to be formatted
# TODO: - column_name: name of the column that should be checked for a value
# TODO: - color: keyworded argument, True if the cell should be colored, False if color should be removed
# TODO: - out_file_suffix: A suffix that will be added to the XLSX file upon saving
# TODO: It should iterate all cells in a row, then find out if the current column is the last column in the file
# TODO: Then, if the cell value is False, the background color of the entire row should be changes.
# TODO: To do this, we will have to iterate over all cells multiple times!
# TODO: We need to find out the number of the column we want to check in each row, the rows that should be colored and
# TODO: then we need to apply the coloring!

def toggle_non_translatable_background_color(xlsx_path, column_name, color=True, out_file_suffix='_bg_color_changed'):
    """
    Changes the background color of cells of a row in an XLSX file, if the value of the last cell in that row is True.
    :param xlsx_path: path of XLSX file to be formatted
    :param column_name: header name of a column. If the value in this column is False, the respective row will be colored
    :param color: if True, cells will be colored red. If False, cell color will be removed
    :param out_file_suffix: Suffix added to the output file before saving
    :return: None
    """
    workbook = openpyxl.load_workbook(xlsx_path)
    colored_rows = list()
    target_col = None
    for worksheet in workbook:
        rows = worksheet.rows
        header_row = worksheet[1]
        for cell in header_row:
            if cell.value == column_name:
                target_col = cell.column
        for row_index, row in enumerate(rows):
            for cell in row:
                if cell.column == target_col and not cell.value:
                    colored_rows.append(cell.row)
        for row_index in colored_rows:
            row = worksheet[row_index]
            for cell in row:
                if color:
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
                else:
                    cell.fill = PatternFill(fill_type=None)
    hidden_out_dir = os.path.dirname(xlsx_path)
    hidden_out_file = Path(xlsx_path).stem + out_file_suffix + '.xlsx'
    hidden_out_path = os.path.join(hidden_out_dir, hidden_out_file)
    workbook.save(hidden_out_path)


toggle_non_translatable_background_color(sample_xlsx_path, 'Translate')


# *** CHANGE FONT COLOR *** #
# Change the text color of non translatable values in a cell

def toggle_font_color(xlsx_path, color=True, out_file_suffix='_font_color_changed'):
    """
    Changes the font color of cells of a row in an XLSX file.
    :param xlsx_path: path of XLSX file to be formatted
    :param color: if True, cell text font will be colored red. If False, cell text font color will be removed
    :param out_file_suffix: Suffix added to the output file before saving
    :return: None
    """
    workbook = openpyxl.load_workbook(xlsx_path)
    for worksheet in workbook:
        column = worksheet['D']
        for cell in column:
            if color:
                cell.font = Font(color='FF0000')
            else:
                cell.font = Font(color='000000')
    hidden_out_dir = os.path.dirname(xlsx_path)
    hidden_out_file = Path(xlsx_path).stem + out_file_suffix + '.xlsx'
    hidden_out_path = os.path.join(hidden_out_dir, hidden_out_file)
    workbook.save(hidden_out_path)


toggle_font_color(sample_xlsx_path)


# *** CHANGE NON TRANSLATABLE FONT COLOR *** #
# Change the text color of non translatable values in a cell

# TODO: Openpyxl doesn't allow us to change the color of individual words in a cell
# TODO: However, we can still check the values of individual cells and color them accordingly!
# TODO: We'd like to write a function "toggle_non_translatable_font_color"
# TODO: The function should take the following arguments:
# TODO: - xlsx_path: positional argument, path of the XLSX file to be formatted
# TODO: - color: keyworded argument, True if the cell should be colored, False if color should be removed
# TODO: - out_file_suffix: A suffix that will be added to the XLSX file upon saving
# TODO: It should also take an argument "non_translatables" which is a list of non translatable cell values
# TODO: It should then iterate all cells in a document and check, if they equal any of the values in "non_translatables"
# TODO: If yes, their font color should be changed to red!

def toggle_non_translatable_font_color(xlsx_path, non_translatables=list(), color=True, out_file_suffix='_font_color_changed'):
    """
    Changes the font color of cells of a row in an XLSX file, if the value is part of a list.
    :param non_translatables: list of values that should be colored in the XLSX file
    :param xlsx_path: path of XLSX file to be formatted
    :param color: if True, cell text font will be colored red. If False, cell text font color will be removed
    :param out_file_suffix: Suffix added to the output file before saving
    :return: None
    """
    workbook = openpyxl.load_workbook(xlsx_path)
    for worksheet in workbook:
        rows = worksheet.rows
        for row_index, row in enumerate(rows):
            for cell in row:
                if cell.value in non_translatables:
                    if color:
                        cell.font = Font(color='FF0000')
                    else:
                        cell.font = Font(color='000000')
    hidden_out_dir = os.path.dirname(xlsx_path)
    hidden_out_file = Path(xlsx_path).stem + out_file_suffix + '.xlsx'
    hidden_out_path = os.path.join(hidden_out_dir, hidden_out_file)
    workbook.save(hidden_out_path)


non_translatables = ['ITServices']
# toggle_non_translatable_font_color(sample_xlsx_path, non_translatables, color=False)


# *** ADD DATA VALIDATION TO CELLS *** #
# Add data validation to specific cells in a XLSX document

def toggle_length_validation(xlsx_path, validate=True, length_limit=100, out_file_suffix='_validation_added'):
    """
    Adds length validation to the translatable text in an XLSX file
    :param xlsx_path: path of XLSX file to be formatted
    :param validate: if True, validation will be added. If False, existing validation will be removed
    :param out_file_suffix: Suffix added to the output file before saving
    :return: None
    """
    workbook = openpyxl.load_workbook(xlsx_path)
    for worksheet in workbook:
        if validate:
            cell_validation = DataValidation(type="textLength", operator="lessThanOrEqual", formula1=length_limit)
            cell_validation.add(f'C{worksheet.min_row}:C{worksheet.max_row}')
            worksheet.add_data_validation(cell_validation)
        else:
            worksheet.data_validations = DataValidationList()
    hidden_out_dir = os.path.dirname(xlsx_path)
    hidden_out_file = Path(xlsx_path).stem + out_file_suffix + '.xlsx'
    hidden_out_path = os.path.join(hidden_out_dir, hidden_out_file)
    workbook.save(hidden_out_path)


toggle_length_validation(sample_xlsx_path)


# *** ADD DATA VALIDATION TO TEXT CELLS *** #
# Add data validation to cells containing text values in a XLSX document

# TODO: We can set data validation for a specific column but what if a fixed value is not precise enough?
# TODO: We can write another function "toggle_text_length_validation"
# TODO: The function should take the following arguments:
# TODO: - xlsx_path: positional argument, path of the XLSX file to be formatted
# TODO: - validate: keyworded argument, True if cells should be validated, False if validation should be removed
# TODO: - out_file_suffix: A suffix that will be added to the XLSX file upon saving
# TODO: The function should iterate all cells in a document and check if they contain strings
# TODO: If they do, it should add character length validation to them!
# TODO: Each text cell should only be allowed contain 1.5 times it's current character count

def toggle_text_length_validation(xlsx_path, validate=True, limit=1.5, out_file_suffix='_validation_added'):
    """
    Adds length validation to all text cells in an XLSX file
    :param xlsx_path: path of XLSX file to be formatted
    :param validate: if True, validation will be added. If False, existing validation will be removed
    :param limit: how many characters are allowed in a text cell. E.g. 2 means two times the original character count
    :param out_file_suffix: Suffix added to the output file before saving
    :return: None
    """
    workbook = openpyxl.load_workbook(xlsx_path)
    for worksheet in workbook:
        rows = worksheet.rows
        if validate:
            for row_index, row in enumerate(rows):
                for cell in row:
                    cell_coordinates = cell.coordinate
                    cell_value = cell.value
                    if isinstance(cell_value, str):
                        cell_value = cell.value
                        cell_chars = len(cell_value)
                        cell_limit = round(cell_chars * limit)
                        cell_validation = DataValidation(type="textLength", operator="lessThanOrEqual", formula1=cell_limit)
                        cell_validation.add(cell_coordinates)
                        worksheet.add_data_validation(cell_validation)
        else:
            worksheet.data_validations = DataValidationList()
    hidden_out_dir = os.path.dirname(xlsx_path)
    hidden_out_file = Path(xlsx_path).stem + out_file_suffix + '.xlsx'
    hidden_out_path = os.path.join(hidden_out_dir, hidden_out_file)
    workbook.save(hidden_out_path)


toggle_length_validation(sample_xlsx_path)
