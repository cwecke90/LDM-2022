# *** LEVEL 2 - GET CELL COMMENTS
# Excel files can have comments attached to each cell but CAT tools can't read them as easily
# If we want to include cell comments as context information, we have to add them as new columns!
# To do this, we have to read the comments first.

import os
import re
import openpyxl

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from pathlib import Path

sample_file = os.path.join('files', 'sample_translation.xlsx')


# TASK 1
# Imagine we have a lot of translatable files with comments attached to individual cells
# We want to find those comments, read them and put them into their own column
# E.g. if you find a comment in a cell within column "Source", it should be moved to a column "Source Comment"
# Meaning if "Source Comment" can't be found in the header row, we have to write it there!
# You will have to add and name columns dynamically for this to work!
# Note, that the text in the new column should be on the same row that it is found in
# Write a function "add_comments_as_columns" that gets comments from a cell and writes them to their own column
# - "xlsx_path": for the path to an XLSX file that should be edited
# It should take one keyworded argument:
# - "out_file_suffix": the suffix added to a new XLSX saved by the funxtion. Default value should be "_comments_added"
# It should return nothing
# The function should check every cell in a document if it has a comment
# If it does, the comment should be written to a new column in the same row
# Hint: much like cell.hidden, there's an attribute that contains comments!
# Hint: comment values consist of more than just comment text. You may have to use regex to get what you want!

# TODO: Write your code here! Feel free to delete this comment!

def add_comments_as_columns(xlsx_path, out_file_suffix='_comments_added'):
    """
    Finds comments in all cells of an XLSX file and writes them to a separate column.
    :param xlsx_path: path of XLSX file to be formatted
    :param out_file_suffix: suffix added to the output file name before saving
    :return: None
    """
    workbook = openpyxl.load_workbook(xlsx_path)
    for worksheet in workbook:
        rows = worksheet.rows
        comment_cols = dict()
        for row in rows:
            for cell in row:
                cell_column = get_column_letter(cell.column)
                print(cell_column)
                cell_row = cell.row
                print("----")
                print(cell_row)
                cell_header = worksheet[f'{cell_column}1'].value
                print("----")
                print(cell_header)
                cell_comment = cell.comment
                if not cell_comment:
                    continue
                comment_text = cell_comment.text
                comment_text = re.search(r'Comment:\n\s+(.+)$', comment_text).group(1)
                if cell_header not in comment_cols:
                    current_max_col = get_column_letter(worksheet.max_column+1)
                    worksheet[f'{current_max_col}1'].font = Font(bold=True)
                    worksheet[f'{current_max_col}1'] = f'{cell_header} comment'
                    comment_cols[cell_header] = current_max_col
                worksheet[f'{comment_cols[cell_header]}{cell_row}'] = comment_text
    # hidden_out_dir = os.path.dirname(xlsx_path)
    # hidden_out_file = Path(xlsx_path).stem + out_file_suffix + '.xlsx'
    # hidden_out_path = os.path.join(hidden_out_dir, hidden_out_file)
    # workbook.save(hidden_out_path)


# TASK 2
# Make sure your function works by using it!
# Call your function "add_comments_as_columns" with "sample_file" as the only argument.
# If everything works, you should find a new file in the same directory as "sample_file"
# When you start working on this homework you will find a file already there to help you visualize the output!

# TODO: Write your code here! Feel free to delete this comment!

add_comments_as_columns(sample_file)