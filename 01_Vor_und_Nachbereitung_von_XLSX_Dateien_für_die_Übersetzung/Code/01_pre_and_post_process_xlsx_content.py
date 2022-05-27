# *** PRE AND POST PROCESSING XLSX CONTENT *** #
# Automatically correcting the content of XLSX files for translation

import os
import openpyxl

import pandas as pd
import numpy as np

from pathlib import Path
from openpyxl.utils import get_column_letter

sample_xlsx_dir = 'files'
sample_xlsx_file = 'DPDHL-Chatbot-Export.xlsx'
sample_xlsx_path = os.path.join(sample_xlsx_dir, sample_xlsx_file)


# *** REFORMATING XLSX CONTENTS *** #
# Editing the contents of XLSX files to prepare them for translation.

# TODO: To kick things off, we'll have a look at some content exported from the DPDHL chatbot system
# TODO: We will take a look at it together and figure out, how it needs to be translated easily
# TODO: To this end, we'll have to take a good look at the file first before writing any code!
# TODO: Then we have to come up with code that can process this file format specifically
# TODO: Note that we must also  get the translated content back into the original format!

xlsx_content_df = pd.read_excel(sample_xlsx_path)


def pre_process_chatbot_export(original_df, src_lang="en", trg_lang="jp"):
    """
    Reformats DPDHL chatbot export files and prepares them for translation.
    :param original_df: a data frame with the contents of a chatbot export
    :param src_lang: the source language that should be used for translation
    :param trg_lang: the language that content should be translated to
    :return: the data to be translated
    """
    # Create the df that should contain the formatted data after changing the original df so all column names are
    # applied!
    prev_id = None
    columns = ['Intent', 'en', 'de', 'es']
    current_row = pd.Series(index=columns)
    result_df = pd.DataFrame(columns=columns)
    iterations = 0

    for index, row in original_df.iterrows():
        if not iterations:
            iterations += 1
        id = row['Intent']
        language = row['Language']
        text = row['Training phrases']
        if id != prev_id:
            result_df = result_df.append(current_row, ignore_index=True)
            current_row = pd.Series(index=columns)
            current_row['Intent'] = id
            prev_id = id
        else:
            current_row[language] = text
    return result_df


def post_process_chatbot_export(original_df, translated_df):
    """
    Reformats translated DPDHL chatbot export files after translation.
    :param original_df: a data frame with the original contents of a chatbot export
    :param translated_df: a data frame with the translated contents of a chatbot export
    :return: the restored original data format with translations
    """
    return None


xlsx_pre_processed_df = pre_process_chatbot_export(xlsx_content_df)


# *** FORMATTING FILES WITH OPENPYXL *** #
# Changing file formats with using openpyxl before saving

def save_processed_df(processed_df, out_path):
    """
    Saves a processed data frame to disk as an XLSX file.
    :param processed_df: a data frame to save to disk
    :param out_path: the path of the new XLSX file
    :return: None
    """
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:

        sheet_name = 'Content'
        processed_df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

        workbook = writer.book
        content_worksheet = workbook[sheet_name]
        column_width = 40
        # Set column width
        for idx, col in enumerate(content_worksheet.columns, 1):
            content_worksheet.column_dimensions[get_column_letter(idx)].width = column_width


pre_process_out_dir_name = sample_xlsx_dir
pre_process_out_file_name = Path(sample_xlsx_file).stem
pre_process_out_path = os.path.join(pre_process_out_dir_name, f'{pre_process_out_file_name}_pre_processed.xlsx')
save_processed_df(xlsx_pre_processed_df, pre_process_out_path)


# *** HIDING COLUMNS AND ROWS *** #
# Using openpyxl to hide certain rows and columns in a file

def toggle_column_visibility(xlsx_path, column_letter, hide=True, out_file_suffix='_cols_hidden'):
    """
    Changes the visibility of the Intent ID column in a DPDHL chatbot export file
    :param xlsx_path: path to an XLSX file
    :param column_letter: letter of column to be hidden
    :param hide: True if the column should be hidden, False if it should be displayed
    :param out_file_suffix: Suffix added to the output file. If no suffix is added, it will be overwritten
    :return: None
    """
    workbook = openpyxl.load_workbook(xlsx_path)

    for worksheet in workbook:
        worksheet.column_dimensions[column_letter].hidden = hide

    hidden_out_dir = os.path.dirname(xlsx_path)
    hidden_out_file = Path(xlsx_path).stem + out_file_suffix + '.xlsx'
    hidden_out_path = os.path.join(hidden_out_dir, hidden_out_file)
    workbook.save(hidden_out_path)

toggle_column_visibility(pre_process_out_path, 'A')


# *** HIDING COLUMNS BY COLUMN NAME *** #
# Using openpyxl to hide certain rows and columns in a file

# TODO: We can now hide specific columns in an XLSX file, but what if we want to specify the name of a column to hide?
# TODO: Openpyxl can't easily do this since it can only go by column letters or ids
# TODO: We can however still write a function, that can hide a named column in any XLSX file!
# TODO: We'd like to write a function "change_column_visibility_by_name"
# TODO: - xlsx_path: positional argument, path of the XLSX file to be formatted
# TODO: - column_name: name of the column that should be hidden
# TODO: - hide: keyworded argument, True if cells should be hidden, False if they should be displayed
# TODO: - out_file_suffix: A suffix that will be added to the XLSX file upon saving
# TODO: First we must acquire the workbook object of a file and each worksheet object in it
# TODO: From each worksheet, acquire the header row and check each cell value in it
# TODO: If the cell value equals the name of the column that should be hidden, get the column letter of that cell
# TODO: Once that column name is found, the column can be hidden!
# TODO: Use this logic to hide the "Origin" column in "files/post_editing_samples.xlsx"
# TODO: The column contains information about how a translation was made
# TODO: We might want to hide this information from a translator to avoid it influencing the post editing results!

def toggle_column_visibility_by_name(xlsx_path, column_name, hide=True, out_file_suffix='_cols_hidden'):
    """
    Changes the visibility of a column with a given name in the header row in any XLSX file
    :param xlsx_path: path to an XLSX file
    :param column_name: name of the column that should be hidden
    :param hide: True if the column should be hidden, False if it should be displayed
    :param out_file_suffix: Suffix added to the output file. If no suffix is added, it will be overwritten
    :return: None
    """
    workbook = openpyxl.load_workbook(xlsx_path)
    target_col = None

    for worksheet in workbook:
        header_row = worksheet[1]
        for cell in header_row:
            cell_value = cell.value
            cell_column = cell.column_letter
            if column_name == cell_value:
                target_col = cell_column
        if target_col:
            worksheet.column_dimensions[target_col].hidden = hide
    hidden_out_dir = os.path.dirname(xlsx_path)
    hidden_out_file = Path(xlsx_path).stem + out_file_suffix + '.xlsx'
    hidden_out_path = os.path.join(hidden_out_dir, hidden_out_file)
    workbook.save(hidden_out_path)


toggle_column_visibility_by_name(pre_process_out_path, 'Intent', out_file_suffix='_intent_hidden')
